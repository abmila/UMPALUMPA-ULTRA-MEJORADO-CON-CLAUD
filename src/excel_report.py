# -*- coding: utf-8 -*-
"""
src/excel_report.py — Generador del reporte Excel ejecutivo

Construye el Excel final profesional con todas las hojas del sistema.
Usa xlsxwriter para formato avanzado (colores, columnas auto-fit, etc.)
con openpyxl como fallback.

Hojas planeadas (en orden):
  1.  DASHBOARD_EJECUTIVO      — Resumen de todo en una página
  2.  RESUMEN_MERCADO          — Estado actual del mercado
  3.  REGIMEN_MACRO            — Régimen macro y sectores favorecidos
  4.  CURVA_BONOS              — Curva de rendimientos UST
  5.  TASAS_EUA_MX             — Tasas de interés EUA y México
  6.  INFLACION_EUA_MX         — Inflación comparada
  7.  COMMODITIES_FX           — Petróleo, oro, plata, USD/MXN, DXY
  8.  QUE_VA_A_SUBIR           — Activos con señal alcista
  9.  QUE_VA_A_BAJAR           — Activos con señal bajista
  10. SECTORES_FAVORECIDOS     — Por régimen macro + señales
  11. SECTORES_PRESIONADOS     — Por régimen macro + señales
  12. INCERTIDUMBRE            — Score de incertidumbre por activo
  13. RANKING_ACTIVOS          — Ranking final con score compuesto
  14. DCF_VALUATION            — Valuaciones DCF detalladas
  15. SENSIBILIDAD_DCF         — Análisis de sensibilidad WACC × g
  16. RATIOS_FINANCIEROS       — Ratios contables por empresa
  17. CALIDAD_FINANCIERA       — Salud y flags por empresa
  18. PORTAFOLIO_OPTIMO        — Portafolio Max Sharpe con pesos
  19. PORTAFOLIO_CANTIDADES    — Cantidades enteras por presupuesto
  20. SUPUESTOS_Y_FUENTES      — Documentación de supuestos y fuentes
  21. WARNINGS_Y_LIMITACIONES  — Alertas y limitaciones del análisis
  22. RAW_DATA_SUMMARY         — Resumen de datos descargados

Estado: STUB — implementación completa en Fase 5
"""

import logging
import pathlib
from datetime import datetime
from typing import Any, Dict, List, Optional

import pandas as pd

log = logging.getLogger(__name__)


def create_excel_report(
    output_data: Dict[str, Any],
    output_path: str = "outputs/resultado_analisis_financiero.xlsx",
) -> str:
    """Genera el reporte Excel ejecutivo completo.

    Args:
        output_data: Dict con todos los datos a incluir. Claves esperadas:
          - 'dcf_results': DataFrame de valuaciones
          - 'health_results': DataFrame de salud financiera
          - 'macro_data': Dict de series macro
          - 'regime': Dict de régimen macro
          - 'probs': DataFrame de probabilidades
          - 'ranking': DataFrame de ranking
          - 'portfolio': Dict de portafolio óptimo
          - 'config': Configuración usada
          - 'warnings': Lista de warnings generados
          - 'run_timestamp': datetime del análisis
          Claves faltantes generan hojas con mensaje de "pendiente".

        output_path: Ruta donde guardar el Excel.

    Returns:
        Ruta absoluta del archivo generado.

    TODO (Fase 5): Implementar.
    """
    raise NotImplementedError("Fase 5: implementar generación completa del Excel")


def write_stub_excel(
    output_path: str = "outputs/resultado_analisis_financiero.xlsx",
    run_info: Optional[Dict] = None,
) -> str:
    """Genera un Excel mínimo de prueba para verificar que el sistema funciona.

    Solo para uso en Fase 1 — confirma que xlsxwriter/openpyxl funcionan.

    Args:
        output_path: Ruta de salida.
        run_info: Dict con info básica del run (timestamp, tickers, etc.).

    Returns:
        Ruta del archivo generado.

    TODO (Fase 2): Expandir con datos reales de macro.
    """
    import os

    path = pathlib.Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    run_info = run_info or {}
    timestamp = run_info.get("timestamp", datetime.now().strftime("%Y-%m-%d %H:%M"))
    tickers = run_info.get("tickers", [])
    phases_done = run_info.get("phases_done", ["Fase 1"])

    try:
        import xlsxwriter

        workbook = xlsxwriter.Workbook(str(path))
        worksheet = workbook.add_worksheet("ESTADO_DEL_SISTEMA")

        bold = workbook.add_format({"bold": True, "font_size": 12})
        title_fmt = workbook.add_format(
            {"bold": True, "font_size": 14, "bg_color": "#1F3864", "font_color": "white"}
        )
        ok_fmt = workbook.add_format({"bg_color": "#C6EFCE", "font_color": "#276221"})
        pending_fmt = workbook.add_format({"bg_color": "#FFEB9C", "font_color": "#9C6500"})

        worksheet.set_column("A:A", 35)
        worksheet.set_column("B:B", 50)

        worksheet.write(0, 0, "Sistema Cuantitativo Unificado v0.1", title_fmt)
        worksheet.write(0, 1, f"Generado: {timestamp}", title_fmt)
        worksheet.write(2, 0, "Tickers configurados:", bold)
        worksheet.write(2, 1, ", ".join(tickers) if tickers else "(sin tickers)")
        worksheet.write(3, 0, "Fases completadas:", bold)
        worksheet.write(3, 1, ", ".join(phases_done))

        worksheet.write(5, 0, "Hoja", bold)
        worksheet.write(5, 1, "Estado", bold)

        sheets_status = [
            ("DASHBOARD_EJECUTIVO", "Pendiente — Fase 5"),
            ("RESUMEN_MERCADO", "Pendiente — Fase 2"),
            ("REGIMEN_MACRO", "Pendiente — Fase 4"),
            ("CURVA_BONOS", "Pendiente — Fase 2"),
            ("TASAS_EUA_MX", "Pendiente — Fase 2"),
            ("INFLACION_EUA_MX", "Pendiente — Fase 2"),
            ("COMMODITIES_FX", "Pendiente — Fase 2"),
            ("QUE_VA_A_SUBIR", "Pendiente — Fase 4"),
            ("QUE_VA_A_BAJAR", "Pendiente — Fase 4"),
            ("SECTORES_FAVORECIDOS", "Pendiente — Fase 4"),
            ("SECTORES_PRESIONADOS", "Pendiente — Fase 4"),
            ("INCERTIDUMBRE", "Pendiente — Fase 4"),
            ("RANKING_ACTIVOS", "Pendiente — Fase 4"),
            ("DCF_VALUATION", "Pendiente — Fase 3"),
            ("SENSIBILIDAD_DCF", "Pendiente — Fase 3"),
            ("RATIOS_FINANCIEROS", "Pendiente — Fase 3"),
            ("CALIDAD_FINANCIERA", "Pendiente — Fase 3"),
            ("PORTAFOLIO_OPTIMO", "Pendiente — Fase 3"),
            ("PORTAFOLIO_CANTIDADES", "Pendiente — Fase 3"),
            ("SUPUESTOS_Y_FUENTES", "En diseño — Fase 5"),
            ("WARNINGS_Y_LIMITACIONES", "En diseño — Fase 5"),
            ("RAW_DATA_SUMMARY", "Pendiente — Fase 2"),
        ]

        for i, (sheet_name, status) in enumerate(sheets_status):
            fmt = ok_fmt if "Completada" in status else pending_fmt
            worksheet.write(6 + i, 0, sheet_name, fmt)
            worksheet.write(6 + i, 1, status, fmt)

        workbook.close()
        log.info("Excel de estado generado: %s", path)

    except ImportError:
        # Fallback con openpyxl si xlsxwriter no está disponible
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "ESTADO_DEL_SISTEMA"
            ws["A1"] = f"Sistema Cuantitativo Unificado v0.1 — {timestamp}"
            ws["A2"] = f"Tickers: {', '.join(tickers) if tickers else '(sin tickers)'}"
            wb.save(str(path))
        except Exception as exc:
            log.error("No se pudo generar Excel: %s", exc)
            return ""

    return str(path.resolve())
